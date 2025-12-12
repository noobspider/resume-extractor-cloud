import os, re, json
import fitz
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
from docx import Document
import phonenumbers
import openpyxl
from io import BytesIO
from openai import OpenAI
from dotenv import load_dotenv
load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
client = OpenAI(api_key=OPENAI_API_KEY)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[A-Za-z]{2,}")

def extract_text_from_pdf(path):
    text = ""
    try:
        doc = fitz.open(path)
        for p in doc:
            text += p.get_text()
        doc.close()
    except Exception:
        text = ""
    if len(text.strip()) < 80:
        try:
            images = convert_from_path(path)
            parts = []
            for img in images:
                parts.append(pytesseract.image_to_string(img))
            if parts:
                text = "\\n".join(parts)
        except Exception:
            pass
    return text

def extract_text_from_docx(path):
    try:
        doc = Document(path)
        return "\\n".join(p.text for p in doc.paragraphs)
    except Exception:
        return ""

def extract_text_from_txt(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception:
        return ""

def extract_text_from_image(path):
    try:
        return pytesseract.image_to_string(Image.open(path))
    except Exception:
        return ""

def extract_text(path):
    _, ext = os.path.splitext(path.lower())
    if ext == ".pdf":
        return extract_text_from_pdf(path)
    if ext == ".docx":
        return extract_text_from_docx(path)
    if ext == ".txt":
        return extract_text_from_txt(path)
    if ext in {".jpg", ".jpeg", ".png"}:
        return extract_text_from_image(path)
    return ""

def detect_emails(text):
    return list(dict.fromkeys(EMAIL_RE.findall(text)))

def detect_phones(text, region="IN"):
    phones = []
    try:
        for m in phonenumbers.PhoneNumberMatcher(text, region):
            phones.append(phonenumbers.format_number(m.number, phonenumbers.PhoneNumberFormat.E164))
    except Exception:
        pass
    return list(dict.fromkeys(phones))

def llm_extract_structured(text):
    max_chars = 25000
    t = text[:max_chars]
    prompt = f\"\"\"You are a resume parser. Return ONLY valid JSON with keys:
Name, Email, Phone, Experience, Skills, Education, Current Company, Location, Summary

Rules:
- Name: candidate full name string or empty
- Email: single email string or empty
- Phone: single phone string (E.164 preferred) or empty
- Experience: total years text (e.g., "5 years") or empty
- Skills: array of skill strings
- Education: array of education lines
- Current Company: string or empty
- Location: string or empty
- Summary: short 1-2 sentence summary

Resume text:
\"\"\"{t}
\"\"\"
\"\"\"
    try:
        resp = client.responses.create(model="gpt-4.1-mini", input=prompt, max_output_tokens=800)
        raw = getattr(resp, "output_text", "") or ""
        s = raw.find("{"); e = raw.rfind("}")
        if s != -1 and e != -1 and e > s:
            data = json.loads(raw[s:e+1])
            for k in ["Name","Email","Phone","Experience","Skills","Education","Current Company","Location","Summary"]:
                if k not in data:
                    data[k] = "" if k not in ["Skills","Education"] else []
            return data
    except Exception:
        pass
    return {"Name":"","Email":"","Phone":"","Experience":"","Skills":[],"Education":[],"Current Company":"","Location":"","Summary":""}

def build_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumes"
    headers = ["Name","Email","Phone","Experience","Skills","Education","Current Company","Location","Summary"]
    ws.append(headers)
    # styling
    from openpyxl.styles import Font, Alignment
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    import openpyxl.utils
    widths = [25,30,18,12,40,40,25,20,60]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for r in records:
        skills = ", ".join(r.get("Skills") or [])
        edu = "; ".join(r.get("Education") or [])
        ws.append([r.get("Name",""), r.get("Email",""), r.get("Phone",""), r.get("Experience",""), skills, edu, r.get("Current Company",""), r.get("Location",""), r.get("Summary","")])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio